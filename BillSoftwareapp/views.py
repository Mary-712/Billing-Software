from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render,redirect, reverse
from django.db.models import F
from django.http import Http404
from io import BytesIO
from xhtml2pdf import pisa
from django.core.mail import send_mail,EmailMultiAlternatives,EmailMessage
from django.template.loader import get_template
from django.conf import settings
import json
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from openpyxl import load_workbook
from openpyxl import Workbook
from django.http import JsonResponse, HttpResponseNotFound
from django.contrib.auth.models import User, auth
from datetime import datetime,date
from . models import *
from django.contrib import messages
from django.utils.crypto import get_random_string
from BillSoftwareapp .models import ItemModel,party,staff_details,company,PurchaseBill,PurchaseBillItem
from django.http import JsonResponse
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from django.db import transaction
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db.models import Max
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def home(request):
  return render(request, 'home.html')


def about(request):
  return render(request, 'about.html')

def contact(request):
  return render(request, 'contact.html')


def service(request):
  return render(request, 'service.html')


def homepage(request):
  return render(request, 'companyhome.html')


def staffhome(request):
  return render(request, 'staffhome.html')

def register(request):
  return render(request, 'register.html')

def registercompany(request):
  return render(request, 'registercompany.html')

def registerstaff(request):
  return render(request, 'registerstaff.html')

def login(request):
  return render(request, 'login.html')

def logout(request):
    auth.logout(request)
    return redirect('/')

def base(request):
  return render(request, 'base.html')

def registeruser(request):
    if request.method == 'POST':
        first_name = request.POST['fname']
        last_name = request.POST['lname']
        user_name = request.POST['username']
        email_id = request.POST['email']
        mobile = request.POST['phoneno']
        passw = request.POST['pass']
        c_passw = request.POST['re_pass']
        profile_pic=request.FILES.get('image')

        if passw != c_passw:
            messages.error(request, 'Passwords do not match')
            return redirect('register')

        if User.objects.filter(username=user_name).exists():
            messages.error(request, 'Sorry, Username already exists')
            return redirect('register')

        if User.objects.filter(email=email_id).exists():
            messages.error(request, 'Sorry, Email already exists')
            return redirect('register')

        # If everything is okay, save the data
        user_data = User.objects.create_user(
            first_name=first_name,
            last_name=last_name,
            username=user_name,
            email=email_id,
            password=passw
        )
        user_data.save()

        data = User.objects.get(id=user_data.id)
        cust_data = company(contact=mobile,profile_pic=profile_pic, user=data)
        cust_data.save()

        demo_staff=staff_details(company=cust_data,
                                   email=email_id,
                                   position='company',
                                   user_name=user_name,
                                   password=passw,
                                   contact=mobile)
        demo_staff.save()

        # messages.success(request, 'Registration successful')
        return redirect('registercompany')

    return render(request, 'register.html')
  
def add_company(request):
  
  if request.method == 'POST':
    email=request.POST['email']
    user=User.objects.get(email=email)
    
    c =company.objects.get(user = user)
    c.company_name=request.POST['companyname']

    c.address=request.POST['address']
    c.city=request.POST['city']
    c.state=request.POST['state']
    c.country=request.POST['country']
    c.pincode=request.POST['pincode']
    c.pan_number=request.POST['pannumber']
    c.gst_type=request.POST['gsttype']
    c.gst_no=request.POST['gstno']

    code=get_random_string(length=6)
    if company.objects.filter(Company_code = code).exists():
       code2=get_random_string(length=6)
       c.Company_code=code2
    else:
      c.Company_code=code
   
    c.save()

    staff = staff_details.objects.get(email=email,position='company',company=c)
    staff.first_name = request.POST['companyname']
    staff.last_name = ''
    staff.save()

    return redirect('login')  
  return render(request,'registercompany.html') 

def staff_registraction(request):
  if request.method == 'POST':
    fn=request.POST['fname']
    ln=request.POST['lname']
    email=request.POST['email']
    un=request.POST['username']
    ph=request.POST['phoneno']
    pas=request.POST['pass']
    code=request.POST['companycode']

    if company.objects.filter(Company_code=code).exists():
      com=company.objects.get(Company_code=code)
    else:
        messages.info(request, 'Sorry, Company code is Invalide')
        return redirect('registerstaff')
    img=request.FILES.get('image')

    if staff_details.objects.filter(user_name=un).exists():
      messages.info(request, 'Sorry, Username already exists')
      return redirect('registerstaff')
    elif staff_details.objects.filter(email=email).exists():
      messages.info(request, 'Sorry, Email already exists')
      return redirect('registerstaff')
    else:
      
      staff=staff_details(first_name=fn,last_name=ln,email=email,user_name=un,contact=ph,password=pas,img=img,company=com)
      staff.save()
      return redirect('login')

  else:
    print(" error")
    return redirect('registerstaff')
  

  
def loginurl(request):
  if request.method == 'POST':
    user_name = request.POST['username']
    passw = request.POST['pass']
    
    log_user = auth.authenticate(username = user_name,
                                  password = passw)
    
    if log_user is not None:
      auth.login(request, log_user)
        
    if staff_details.objects.filter(user_name=user_name,password=passw,position='company').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='company') 

      request.session["staff_id"]=data.id
      if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
          staff_id = request.session['staff_id']
          print(staff_id)
 
        return redirect('homepage')  

    if staff_details.objects.filter(user_name=user_name,password=passw,position='staff').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='staff')   

      request.session["staff_id"]=data.id
      if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
          staff_id = request.session['staff_id']
          print(staff_id)
 
          return redirect('staffhome')  
    else:
      messages.info(request, 'Invalid Username or Password. Try Again.')
      return redirect('login')  
  else:  
   return redirect('login')   
  
def profile(request):
 
  return render(request, 'profile.html')

def first_page(request):
 
  return redirect('homepage')

def staffprofile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context = {
              'staff' : staff

          }
  return render(request,'profilestaff.html',context)

      
      
      
      
      
      
def show_purchasebill(request,id):
    purchase_bill = get_object_or_404(PurchaseBill, id=id)

    # Fetch other related data if needed
    pitm = PurchaseBillItem.objects.filter(purchasebill=purchase_bill)

    # Calculate discount and other necessary data
    dis = sum(int(itm.discount) for itm in pitm)
    itm_len = len(pitm)

    # Pass the data to the template
    context = {'purchase_bill': purchase_bill, 'pitm': pitm, 'itm_len': itm_len, 'dis': dis}
    return render(request, 'purchasebilldetails.html', context)


def view_purchasebill(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  pbill = PurchaseBill.objects.filter(company=cmp)
  
  if not pbill:
    context = {'staff':staff}
    return render(request,'purchasebillempty.html',context)
  
  context = {'staff':staff,'pbill':pbill}
  return render(request,'purchasebilllist.html',context)


def add_purchasebill(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")

  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  pbill = PurchaseBill.objects.filter(company=cmp)

  
  
  
  last_bill = PurchaseBill.objects.filter(company=cmp).last()

  if last_bill:
    bill_no = last_bill.tot_bill_no + 1 
  else:
    bill_no = 1

  item = ItemModel.objects.filter(company=cmp,user=cmp.user)

  context = {'staff':staff, 'cust':cust, 'cmp':cmp,'bill_no':bill_no, 'tod':tod, 'item':item, }
  return render(request,'purchasebilladd.html',context)


def save_item(request):
    if request.method == 'POST':
        item_name = request.POST.get('item_name')
        hsn = request.POST.get('hsn')
        qty = request.POST.get('qty')
        tax_ref = request.POST.get('taxref')
        intra_st = request.POST.get('intra_st')
        inter_st = request.POST.get('inter_st')
        sale_price = request.POST.get('saleprice')
        purchase_price = request.POST.get('purprice')

        # Perform any additional validation or processing here

        # Create an instance of your model and save it to the database
        item = ItemModel(
            item_name=item_name,
            item_hsn=hsn,
            item_current_stock=qty,
            item_taxable=tax_ref,
            item_gst=intra_st,
            item_igst=inter_st,
            item_sale_price=sale_price,
            item_purchase_price=purchase_price
        )
        item.save()

        # Redirect to a success page or return a success message
        return redirect("add_purchasebill")

    # Handle GET requests or any other cases
    return render(request, 'purchasebilladd.html')


def create_purchasebill(request):
  if request.method == 'POST': 
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    part = party.objects.get(id=request.POST.get('customername'))
    pbill = PurchaseBill(party=part, 
                          billno=request.POST.get('bill_no'),
                          billdate=request.POST.get('billdate'),
                          supplyplace =request.POST.get('placosupply'),
                          pay_method=request.POST.get("method"),
                          cheque_no=request.POST.get("cheque_id"),
                          upi_no=request.POST.get("upi_id"),
                          advance = request.POST.get("advance"),
                          balance = request.POST.get("balance"),
                          subtotal=float(request.POST.get('subtotal')),
                          igst = request.POST.get('igst'),
                          cgst = request.POST.get('cgst'),
                          sgst = request.POST.get('sgst'),
                          adjust = request.POST.get("adj"),
                          taxamount = request.POST.get("taxamount"),
                          grandtotal=request.POST.get('grandtotal'),
                          company=cmp,staff=staff)
    pbill.save()
        
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    billno = PurchaseBill.objects.get(billno =pbill.billno,company=cmp)

    if len(product)==len(qty)==len(discount)==len(total):
        mapped=zip(product,qty,discount,total)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          PurchaseBillItem.objects.create(product = itm,qty=ele[1],discount=ele[2],total=ele[3],purchasebill=billno,company=cmp)

    PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=F('tot_bill_no') + 1)
    
    pbill.tot_bill_no = pbill.billno
    pbill.save()

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Created')

    if 'Next' in request.POST:
      return redirect('add_purchasebill')
    
    if "Save" in request.POST:
      return redirect('view_purchasebill')
    
  else:
    return render(request,'purchasebilladd.html')
  
def edit_purchasebill(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
 
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  billprd = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)

 
  bdate = pbill.billdate.strftime("%Y-%m-%d")
  context = {'staff':staff, 'pbill':pbill, 'billprd':billprd,'tod':tod,
             'cust':cust, 'item':item,  'bdate':bdate}
  return render(request,'purchasebilledit.html',context)


def update_purchasebill(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = party.objects.get(id=request.POST.get('customername'))
    pbill = PurchaseBill.objects.get(id=id,company=cmp)
    pbill.party = part
    pbill.billdate = request.POST.get('billdate')
    pbill.supplyplace  = request.POST.get('placosupply')
    pbill.subtotal =float(request.POST.get('subtotal'))
    pbill.grandtotal = request.POST.get('grandtotal')
    pbill.igst = request.POST.get('igst')
    pbill.cgst = request.POST.get('cgst')
    pbill.sgst = request.POST.get('sgst')
    pbill.taxamount = request.POST.get("taxamount")
    pbill.adjust = request.POST.get("adj")
    pbill.pay_method = request.POST.get("method")
    pbill.cheque_no = request.POST.get("cheque_id")
    pbill.upi_no = request.POST.get("upi_id")
    pbill.advance = request.POST.get("advance")
    pbill.balance = request.POST.get("balance")

    pbill.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
    if len(total)==len(discount)==len(qty):
      mapped=zip(product,qty,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        PurchaseBillItem.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],purchasebill=pbill,company=cmp)

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasebill')

  return redirect('view_purchasebill')


def details_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'pbill':pbill,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'purchasebilldetails.html',context)


def history_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  hst= PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp)

  context = {'staff':staff,'hst':hst,'pbill':pbill}
  return render(request,'purchasebillhistory.html',context)


def delete_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(id=id)
  PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
  pbill.delete()
  return redirect('view_purchasebill')


def import_purchase_bill(request):
  if request.method == 'POST' and 'billfile' in request.FILES and 'prdfile' in request.FILES:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(PurchaseBill.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      PurchaseBill.objects.create(party=part,billno=totval,
                                  billdate=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      pbill = PurchaseBill.objects.last()
      if billsheet[4] == 'Cheque':
        pbill.pay_method = 'Cheque'
        pbill.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        pbill.pay_method = 'UPI'
        pbill.upi_no = billsheet[5]
      else:
       
        pbill.save()

      PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=int(prdsheet[2]),company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          PurchaseBillItem.objects.create(purchasebill=pbill,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

          if billsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if billsheet[3]=='State':
            gst = round((taxamount/2),2)
            pbill.sgst=gst
            pbill.cgst=gst
            pbill.igst=0

          else:
            gst=round(taxamount,2)
            pbill.igst=gst
            pbill.cgst=0
            pbill.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pbill.subtotal=round(subtotal,2)
      pbill.taxamount=round(taxamount,2)
      pbill.adjust=round(billsheet[6],2)
      pbill.grandtotal=gtotal
      pbill.advance=round(billsheet[7],2)
      pbill.balance=balance
      pbill.save()

      PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,staff=pbill.staff,company=pbill.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})
  


def billhistory(request):
  pid = request.POST['id']
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(billno=pid,company=cmp)
  hst = PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid})



def savecustomer(request):
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    party_name = request.POST['name']
    email = request.POST['email']
    contact = request.POST['mobile']
    state = request.POST['splystate']
    address = request.POST['baddress']
    gst_type = request.POST['gsttype']
    gst_no = request.POST['gstin']
    current_date = request.POST['partydate']
    openingbalance = request.POST.get('openbalance')
    payment = request.POST.get('paytype')
    creditlimit = request.POST.get('credit_limit')
    End_date = request.POST.get('enddate', None)

    part = party(party_name=party_name, gstin=gst_no, phone_number=contact, gst_type=gst_type, state=state, billing_address=address,
                 email=email, opening_balance=openingbalance, payment=payment, current_date=current_date, End_date=End_date,
                 company=cmp, user=cmp.user)
    part.save()

    # Return the newly created customer's ID and name
    response_data = {'success': True, 'new_customer_id': part.id, 'new_customer_name': part.party_name}
    return JsonResponse(response_data)

def cust_dropdown(request):
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    part = party.objects.filter(company=cmp, user=cmp.user)

    id_list = []
    party_list = []
    for p in part:
        id_list.append(p.id)
        party_list.append(p.party_name)

    return JsonResponse({'id_list': id_list, 'party_list': party_list})
  
def saveitem(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  if not hsn:
    hsn = None

  itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
              company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})


def item_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
  return JsonResponse({'id_list':id_list, 'product_list':product_list})



def add_item(request):
    if request.method == 'POST':
        # Process form data and create a new item
        # Extract data from the request.POST dictionary
        type = request.POST.get('type')
        name = request.POST.get('name')
        hsn = request.POST.get('hsn')
        unit = request.POST.get('unit')
        tax = request.POST.get('tax')
        sprice = request.POST.get('sprice', 0)
        pprice = request.POST.get('pprice', 0)
        gst = request.POST.get('gst', 0)
        igst = request.POST.get('igst', 0)
        sth = request.POST.get('sth')
        extraPrice = request.POST.get('extraPrice')
        date = request.POST.get('date')

        # Handle empty strings for numeric fields
        sprice = 0 if not sprice else sprice
        pprice = 0 if not pprice else pprice
        gst = 0 if not gst else gst
        igst = 0 if not igst else igst

        # Create a new ItemModel instance
        new_item = ItemModel.objects.create(
            item_name=name,
            item_hsn=hsn,
            item_unit=unit,
            item_taxable=tax,
            item_gst=gst,
            item_igst=igst,
            item_sale_price=sprice,
            item_purchase_price=pprice,
            item_stock_in_hand=sth,
            item_at_price=extraPrice,
            item_date=date,
            item_type=type,
        )

        # Get all items to update the dropdown
        items = ItemModel.objects.values('id', 'item_name')

        # Return the newly created item and the updated dropdown
        return JsonResponse({'success': True, 'new_item': {'id': new_item.id, 'name': new_item.item_name}, 'items': list(items)})
    
    return render(request, 'purchasebilladd.html.html')

def custdata(request):
  cid = request.POST['id']
  part = party.objects.get(id=cid)
  phno = part.phone_number
  address = part.billing_address
  pay = part.payment
  bal = part.opening_balance
  return JsonResponse({'phno':phno, 'address':address, 'pay':pay, 'bal':bal})

def history(request):
  return render(request,'purchasebillhistory.html')
def itemdetails(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})

def sharepdftomail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                sid = request.session.get('staff_id')
                staff = staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
                
                pbill = PurchaseBill.objects.get(id=id,company=cmp)
                pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
                context = {'pbill':pbill, 'cmp':cmp,'pitm':pitm}
                template_path = 'purchase_mail.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'DEBIT NOTE - {pbill.id}.pdf'
                subject = f"DEBIT NOTE - {pbill.id}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached DEBIT NOTE - File-{pbill.id}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
                return redirect('details_purchasebill', id=id)

        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('details_purchasebill', id=id)

 




    
